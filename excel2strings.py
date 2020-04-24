from openpyxl import load_workbook

print("Processing ...\n")

wb = load_workbook(filename = 'income.xlsx')

sheet_name_list = [
    "2016",
    "2017",
    "2018"
]

key_list = [] # 主要为记录顺序
map_en = {}
map_tc = {}
map_sc = {}

def trim_key(text):
    return text.strip()
def trim_value(text):
    if text is None:
        return ""
    return text.strip().replace("\"", "\\\"").replace("\n", "\\n").replace("\\n\"", "\"")

with open("output.log", 'w') as file_log:
    for sheet_name in sheet_name_list:
        sheet = wb[sheet_name]
        sheet_name_key = "SHEET_NAME_%s" % sheet.title
        key_list.append(sheet_name_key)
        map_en[sheet_name_key] = sheet.title
        map_tc[sheet_name_key] = sheet.title
        map_sc[sheet_name_key] = sheet.title
        file_log.write("\n")
        file_log.write("---------------------------Sheet:%s----------------------\n" % sheet.title)
        rows = sheet.iter_rows(min_row=2,min_col=2,max_col=5, values_only=True)
        for index, row in enumerate(rows):
            # blank line: 4 cell all None, skip the row
            if (row[0] is None) and (row[1] is None) and (row[2] is None) and (row[3] is None):
                continue
            # blank cell: 1 or 2, or 3 cell value is None, write to log, then skip the row
            elif (row[0] is None) or (row[1] is None) or (row[2] is None) or (row[3] is None):
                file_log.write("\n")
                file_log.write("Sheet Name ->:[%s] - Line number:%d\n" % (sheet.title, index + 2))
                file_log.write("Key -------->:%s\n" % (row[0]))
                file_log.write("Value EN --->:%s\n" % (row[1]))
                file_log.write("Value TC --->:%s\n" % (row[2]))
                file_log.write("Value SC --->:%s\n" % (row[3]))
                file_log.write("\n")
                continue
            # OK line: join in map_xx, waiting to write to xx.strings file.
            else:
                key = trim_key(row[0])
                value_en = trim_value(row[1])
                value_tc = trim_value(row[2])
                value_sc = trim_value(row[3])

                if key_list.count(key) != 0:
                    key_list.remove(key)
                key_list.append(key)
                map_en[key] = value_en
                map_tc[key] = value_tc
                map_sc[key] = value_sc

with open("output_en.strings", 'w') as file_en:
    with open("output_sc.strings", 'w') as file_sc:
        with open("output_tc.strings", 'w') as file_tc:
            for key in key_list:
                if key.startswith("SHEET_NAME"):
                    sheet_name_value_en = map_en[key]
                    file_en.write("\n")
                    file_en.write("/*    %s    */\n" % sheet_name_value_en)
                    file_en.write("\n")

                    sheet_name_value_tc = map_tc[key]
                    file_tc.write("\n")
                    file_tc.write("/*    %s    */\n" % sheet_name_value_tc)
                    file_tc.write("\n")

                    sheet_name_value_sc = map_sc[key]
                    file_sc.write("\n")
                    file_sc.write("/*    %s    */\n" % sheet_name_value_sc)
                    file_sc.write("\n")
                else:
                    value_en = map_en[key]
                    value_tc = map_tc[key]
                    value_sc = map_sc[key]
                    file_en.write("\"%s\" = \"%s\";\n" % (key, value_en))
                    file_sc.write("\"%s\" = \"%s\";\n" % (key, value_sc))
                    file_tc.write("\"%s\" = \"%s\";\n" % (key, value_tc))

print("===============================The End====================================\n")
